import streamlit as st
st.set_page_config(page_title="RMI èˆ‡è¯é‚¦è³‡æ–™è™•ç†æ‡‰ç”¨", page_icon=":bar_chart:", layout="wide")
import os
import pandas as pd
from selenium import webdriver
import shutil
import xml.etree.ElementTree as ET
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import ssl
import re
import time
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from st_on_hover_tabs import on_hover_tabs
import streamlit as st
from PIL import Image


if 'rmi_file_path' not in st.session_state:
    st.session_state['rmi_file_path'] = None
if 'merged_file_path' not in st.session_state:
    st.session_state['merged_file_path'] = None

def setup_paths():
    try:
        global base_path
        base_path = st.text_input("è«‹è¼¸å…¥æ¬²å­˜æ”¾æª”æ¡ˆçš„è³‡æ–™å¤¾è·¯å¾‘", placeholder=r"ä¾‹ï¼šC:\Users\YCChi13\Desktop\è¡çªç¤¦ç”¢ç³»çµ±è³‡æ–™å¤¾")
        today_date = datetime.now().strftime('%Y%m%d')  # ç”Ÿæˆä»Šå¤©çš„æ—¥æœŸï¼Œæ ¼å¼ç‚º YYYYMMDD
        compared_path = os.path.join(base_path, "compared")  # å‰µå»ºä¸€å€‹å­˜å„²æ¯”å°çµæœçš„ç›®éŒ„è·¯å¾‘
        if not os.path.exists(compared_path):  
            os.makedirs(compared_path)
        return base_path, today_date, compared_path  # è¿”å›ä¸‰å€‹è·¯å¾‘/æ—¥æœŸè®Šé‡
    except OSError as e:
        if e.winerror == 123:  
            st.warning("è·¯å¾‘ä¸èƒ½åŒ…å«å¼•è™Ÿ \" æˆ– \' æˆ–å…¶ä»–ç„¡æ•ˆå­—ç¬¦ã€‚")
            st.warning(r"åƒè€ƒæ ¼å¼ï¼šC:\Users\YCChi13\Desktop\è¡çªç¤¦ç”¢ç³»çµ±è³‡æ–™å¤¾")
        else:
            st.error(f"ç™¼ç”ŸéŒ¯èª¤: {e}")
        return None, None, None

def download_and_merge_files(base_path, today_date):
    st.header("çˆ¬å–RMI/åˆä½µè¯é‚¦æª”æ¡ˆ")
    st.markdown("""
    è«‹å‰µå»ºä¸€å€‹æ–°è³‡æ–™å¤¾ï¼Œæ”¾ç½®ä»¥ä¸‹è³‡æ–™ï¼š
    - å°è£å» æª”æ¡ˆ
    - ä¾›æ‡‰å•†æª”æ¡ˆ
    """)

    directory_path = st.text_input("è«‹è¼¸å…¥å‰µå»ºçš„è³‡æ–™å¤¾è·¯å¾‘:", base_path)
    download_button = st.button("é–‹å§‹ä¸‹è¼‰å’Œåˆä½µ")

    if download_button:
        if not directory_path:
            st.warning("è«‹é¸æ“‡è³‡æ–™å¤¾è·¯å¾‘")
        else:
            st.write("é–‹å§‹ä¸‹è¼‰...")
            download_path = os.path.join(base_path, "downloads")
            if not os.path.exists(download_path):
                os.makedirs(download_path)

            # ä¸‹è¼‰ä¸¦è™•ç† RMI è³‡æ–™çš„å‡½æ•¸èª¿ç”¨
            download_and_process_rmi_data(base_path, download_path)

            merged_path = os.path.join(base_path, "merged")
            if not os.path.exists(merged_path):
                os.makedirs(merged_path)

            all_path = os.path.join(base_path, "All")

            # åˆä½µæ–‡ä»¶çš„å‡½æ•¸èª¿ç”¨
            merged_df = process_files(directory_path, merged_path)
            st.write("åˆä½µçš„è³‡æ–™è¡¨ï¼š")
            st.dataframe(merged_df)

            # è¨˜éŒ„æ–‡ä»¶è·¯å¾‘åˆ° session state
            st.session_state['rmi_file_path'] = os.path.join(all_path, f"RMI_All_{today_date}.xlsx")
            st.session_state['merged_file_path'] = os.path.join(merged_path, f"General_merged_{today_date}.xlsx")


def datacleaning(data):
    data = data.dropna(how='all')
    data['Smelter Look-up (*)'] = data['Smelter Look-up (*)'].fillna(data['Smelter Name (1)'])
    data['Smelter Name (1)'] = data['Smelter Name (1)'].fillna(data['Smelter Look-up (*)'])
    data['Smelter Identification Number Input Column'] = data['Smelter Identification Number Input Column'].fillna(data['Smelter Identification'])
    data['Smelter Identification'] = data['Smelter Identification'].fillna(data['Smelter Identification Number Input Column'])
    data.loc[data['Source of Smelter Identification Number'].isna(), 'Source of Smelter Identification Number'] = 'RMI'
    return data

def process_files(directory_path, merged_path):
    files = [os.path.join(directory_path, file) for file in os.listdir(directory_path) if file.endswith('.xlsx')]
    columns_to_keep = [
        'Smelter Identification Number Input Column', 'Metal (*)', 'Smelter Look-up (*)', 'Smelter Name (1)', 
        'Smelter Country (*)', 'Smelter Identification', 'Source of Smelter Identification Number', 
        'Smelter Street ', 'Smelter City', 'Smelter Facility Location: State / Province', 'Smelter Contact Name', 
        'Smelter Contact Email', 'Proposed next steps', 'Name of Mine(s) or if recycled or scrap sourced, enter "recycled" or "scrap"', 
        'Location (Country) of Mine(s) or if recycled or scrap sourced, enter "recycled" or "scrap"', 
        'Does 100% of the smelterâ€™s feedstock originate from recycled or scrap sources?', 'Comments', 'Source Name'
    ]

    dfs = []


    for file in files:
        df_declaration = pd.read_excel(file, sheet_name='Declaration', header=None, usecols="D", skiprows=7, nrows=15)
        df_declaration.columns = ['Value']
        source_name = df_declaration.iloc[0, 0].strip() if pd.notna(df_declaration.iloc[0, 0]) and df_declaration.iloc[0, 0].strip() else 'Subcon'
        
        declaration_data = {'Source Name': source_name}

        df_smelter_list = pd.read_excel(file, sheet_name='Smelter List', header=3, nrows=100)

        for key, value in declaration_data.items():
            df_smelter_list[key] = value

        df_smelter_list = datacleaning(df_smelter_list[columns_to_keep])
        dfs.append(df_smelter_list)

    
    dfs = [df.astype(object) for df in dfs]
    
    merged_df = pd.concat(dfs, ignore_index=True, sort=False)
    merged_df['non_na_count'] = merged_df.notna().sum(axis=1)

    merged_df['Source Name'] = merged_df.groupby('Smelter Identification Number Input Column')['Source Name'].transform(lambda x: ', '.join(x.dropna().unique()))

    merged_df = merged_df.loc[
        merged_df.groupby('Smelter Identification Number Input Column')['non_na_count'].idxmax()
    ]

    merged_df = merged_df.drop(columns=['non_na_count'])
    merged_df = merged_df.dropna(subset=['Smelter Identification Number Input Column'])
    merged_df = merged_df[merged_df['Metal (*)'].isin(["Gold", "Tin", "Tantalum", "Tungsten", "Cobalt"])]
    merged_df = merged_df.sort_values(by='Metal (*)')

    output_excel_file_path = os.path.join(merged_path, f"General_merged_{datetime.now().strftime('%Y%m%d')}.xlsx")
    merged_df.to_excel(output_excel_file_path, index=False)
    
    st.success(f"Generalæª”æ¡ˆåˆä½µå®Œæˆï¼Œå·²å„²å­˜åˆ° {output_excel_file_path}.")
    return merged_df

def download_and_process_rmi_data(base_path, download_path):
    os.environ["WDM_SSL_VERIFY"] = "0"
    ssl._create_default_https_context = ssl._create_unverified_context

    chrome_options = ChromeOptions()
    chrome_options.add_experimental_option("prefs", {
        "download.default_directory": download_path,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    })

    chromedriver_path = os.path.join(base_path, "chromedriver.exe")
    driver = webdriver.Chrome(service=ChromeService(chromedriver_path), options=chrome_options)

    driver.get("https://b5.caspio.com/dp/0c4a3000395c678f4f4c4926a081")
    wait = WebDriverWait(driver, 10)

    try:
        download_csv = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Download Data")))
        download_csv.click()
        time.sleep(5)

        files = sorted(os.listdir(download_path), key=lambda x: os.path.getctime(os.path.join(download_path, x)), reverse=True)
        latest_file = files[0]
        
        if not latest_file.endswith('.xml'):
            st.error(f'ä¸‹è¼‰æœªæˆåŠŸï¼Œæ–‡ä»¶ä¸æ˜¯ XML æ–‡ä»¶: {latest_file}')
            return
        
        # ç›´æ¥å°‡æ–°æ–‡ä»¶ä¿å­˜åˆ° base_path\All ç›®éŒ„ä¸‹
        all_path = os.path.join(base_path, "All")
        if not os.path.exists(all_path):
            os.makedirs(all_path)
        
        new_name = f"All_{os.path.splitext(latest_file)[0]}{os.path.splitext(latest_file)[1]}"
        
        tree = ET.parse(os.path.join(download_path, latest_file))
        root = tree.getroot()
        namespaces = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
        rows = root.findall(".//ss:Row", namespaces=namespaces)

        data = []
        for row in rows:
            cells = row.findall(".//ss:Cell", namespaces=namespaces)
            row_data = [cell.find(".//ss:Data", namespaces=namespaces).text if cell.find(".//ss:Data", namespaces=namespaces) is not None else '' for cell in cells]
            row_data.insert(0, "All")
            data.append(row_data)
        
        df = pd.DataFrame(data)
        excel_file_path = os.path.join(all_path, f"RMI_All_{datetime.now().strftime('%Y%m%d')}.xlsx")
        df.to_excel(excel_file_path, index=False, header=False)
        
        st.success(f"æª”æ¡ˆå·²å„²å­˜åˆ° {excel_file_path}")
    except Exception as e:
        st.error(f'ä¸‹è¼‰éç¨‹ä¸­ç™¼ç”ŸéŒ¯èª¤: {e}')
    finally:
        driver.quit()


def process_smelter_data(rmi_df, merge_df, compared_path, today_date):
    rmi_audit_info = rmi_df.set_index('Smelter ID')[['Last Audit Date', 'Audit Cycle']]

    rmi_ids = set(rmi_df['Smelter ID'])
    merge_ids = set(merge_df['Smelter Identification Number Input Column'])

    unmatched_ids = merge_ids - rmi_ids
    matched_ids = merge_ids.intersection(rmi_ids)

    unmatched_rows = merge_df[merge_df['Smelter Identification Number Input Column'].isin(unmatched_ids)]
    matched_rows = merge_df[merge_df['Smelter Identification Number Input Column'].isin(matched_ids)]

    unmatched_rows = unmatched_rows.copy()
    unmatched_rows['Last Audit Date'] = unmatched_rows['Smelter Identification Number Input Column'].map(rmi_audit_info['Last Audit Date'])
    unmatched_rows['Audit Cycle'] = unmatched_rows['Smelter Identification Number Input Column'].map(rmi_audit_info['Audit Cycle'])

    matched_rows = matched_rows.copy()
    matched_rows['Last Audit Date'] = matched_rows['Smelter Identification Number Input Column'].map(rmi_audit_info['Last Audit Date'])
    matched_rows['Audit Cycle'] = matched_rows['Smelter Identification Number Input Column'].map(rmi_audit_info['Audit Cycle'])

    num_unmatched = len(unmatched_ids)
    unmatched_data = unmatched_rows[['Smelter Identification Number Input Column', 'Source Name']].values.tolist()
    num_matched = len(matched_ids)

    def calculate_due_date(row):
        last_audit_date = row['Last Audit Date']
        audit_cycle = row['Audit Cycle']

        if pd.notnull(last_audit_date) and pd.notnull(audit_cycle):
            last_audit_date = pd.to_datetime(last_audit_date)

            cycle_years = int(re.search(r'\d+', audit_cycle).group())  # æå–å¹´ä»½æ•¸å­—
            due_date = last_audit_date + pd.DateOffset(years=cycle_years)
            return due_date
        else:
            return None

    unmatched_rows['Due Date'] = unmatched_rows.apply(calculate_due_date, axis=1)
    matched_rows['Due Date'] = matched_rows.apply(calculate_due_date, axis=1)

    today = datetime.now()
    days_threshold = 30  

    result_text = ""

    def check_due_date(df, result_text):
        for _, row in df.iterrows():
            if pd.notnull(row['Due Date']):
                days_diff = (row['Due Date'] - today).days
                if 0 <= days_diff <= days_threshold:
                    result_text += f"Smelter ID: {row['Smelter Identification Number Input Column']}, ç…‰è£½å» ï¼š{row['Smelter Name (1)']}, ä¾†æºåç¨±: {row['Source Name']}, åˆ°æœŸæ—¥: {row['Due Date'].strftime('%Y-%m-%d')}\n\n"
        return result_text

    result_text = check_due_date(unmatched_rows, result_text)
    result_text = check_due_date(matched_rows, result_text)

    unmatched_path = os.path.join(compared_path, f"Unmatch_General_RMI_{today_date}.xlsx")
    matched_path = os.path.join(compared_path, f"Match_General_RMI_{today_date}.xlsx")
    unmatched_rows.to_excel(unmatched_path, index=False)
    matched_rows.to_excel(matched_path, index=False)

    return unmatched_rows, matched_rows, unmatched_path, matched_path, result_text, num_unmatched, unmatched_data, num_matched

def create_excel_files(merge_df, compared_path, today_date):
    original_file_path = os.path.join(base_path, "RMI_CMRT_6.4.xlsx")
    output_general_path = os.path.join(compared_path, f"å«ä¾†æºåç¨±_General_RMI_CMRT_6.4_{today_date}.xlsx")
    shutil.copyfile(original_file_path, output_general_path)

    wb = openpyxl.load_workbook(output_general_path)
    ws = wb['Smelter List']

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    for r_idx, row in enumerate(dataframe_to_rows(merge_df, index=False, header=True), start=4):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            if 'A' <= cell.column_letter <= 'S':
                cell.border = thin_border
            else:
                cell.border = None
    wb.save(output_general_path)

    # Creating Winbond versions
    global winbond_path
    winbond_path = os.path.join(base_path, "winbondå…¬ç‰ˆ")
    if not os.path.exists(winbond_path):
        os.makedirs(winbond_path)
    
    output_winbond_path = os.path.join(winbond_path, f"Winbond_RMI_CMRT_6.4_General_{today_date}.xlsx")
    shutil.copyfile(output_general_path, output_winbond_path)
    wb_winbond = openpyxl.load_workbook(output_winbond_path)
    ws_winbond = wb_winbond['Smelter List']
    ws_winbond.column_dimensions['R'].hidden = True
    ws_winbond.column_dimensions['S'].hidden = True
    wb_winbond.save(output_winbond_path)

    # Creating KGD versions
    KGD_path = os.path.join(winbond_path, f"Winbond_RMI_CMRT_6.4_KGD_{today_date}.xlsx")
    shutil.copyfile(original_file_path, KGD_path)
    wb_winbond = openpyxl.load_workbook(KGD_path)
    ws_winbond = wb_winbond['Smelter List']
    for row in ws_winbond.iter_rows(min_row=4, max_row=ws_winbond.max_row, max_col=ws_winbond.max_column):
        for cell in row:
            cell.value = None
    KGD_df = merge_df[merge_df['Metal (*)'].isin(['Tantalum', 'Tungsten'])]
    for r_idx, row in enumerate(dataframe_to_rows(KGD_df, index=False, header=True), start=4):
        for c_idx, value in enumerate(row, start=1):
            ws_winbond.cell(row=r_idx, column=c_idx, value=value)
    wb_winbond.save(KGD_path)

    # Creating KGD_RDL versions
    KGD_RDL_path = os.path.join(winbond_path, f"Winbond_RMI_CMRT_6.4_KGD_RDL_{today_date}.xlsx")
    shutil.copyfile(original_file_path, KGD_RDL_path)
    wb_winbond = openpyxl.load_workbook(KGD_RDL_path)
    ws_winbond = wb_winbond['Smelter List']
    for row in ws_winbond.iter_rows(min_row=4, max_row=ws_winbond.max_row, max_col=ws_winbond.max_column):
        for cell in row:
            cell.value = None
    KGD_RDL_df = merge_df[merge_df['Metal (*)'].isin(['Gold', 'Tantalum', 'Tungsten'])]
    for r_idx, row in enumerate(dataframe_to_rows(KGD_RDL_df, index=False, header=True), start=4):
        for c_idx, value in enumerate(row, start=1):
            ws_winbond.cell(row=r_idx, column=c_idx, value=value)
    wb_winbond.save(KGD_RDL_path)

    # Creating WLCSP versions
    WLCSP_path = os.path.join(winbond_path, f"Winbond_RMI_CMRT_6.4_WLCSP_{today_date}.xlsx")
    shutil.copyfile(original_file_path, WLCSP_path)
    wb_winbond = openpyxl.load_workbook(WLCSP_path)
    ws_winbond = wb_winbond['Smelter List']
    for row in ws_winbond.iter_rows(min_row=4, max_row=ws_winbond.max_row, max_col=ws_winbond.max_column):
        for cell in row:
            cell.value = None
    WLCSP_df = merge_df[merge_df['Metal (*)'].isin(['Tin', 'Tantalum', 'Tungsten'])]
    for r_idx, row in enumerate(dataframe_to_rows(WLCSP_df, index=False, header=True), start=4):
        for c_idx, value in enumerate(row, start=1):
            ws_winbond.cell(row=r_idx, column=c_idx, value=value)
    wb_winbond.save(WLCSP_path)

    return output_general_path, winbond_path

def display_results(num_unmatched, unmatched_path, output_general_path, winbond_path, result_text, num_matched, unmatched_data):

    st.subheader("ğŸš¨ Audit Dateåˆ°æœŸæé†’ï¼ˆè¿‘30æ—¥ï¼‰")
    if result_text:
        st.write(result_text)
    else:
        st.write("æ‰€æœ‰æª”æ¡ˆå·²æˆåŠŸç”Ÿæˆä¸¦å„²å­˜ï¼Œç„¡æ¥è¿‘åˆ°æœŸçš„è¨˜éŒ„ã€‚")

    st.subheader("æ¯”å°çµæœ")
    st.markdown(f"<p style='color:red; font-weight:bold;'>ä¸Šå‚³æª”æ¡ˆä¸ç¬¦RMIçš„ Smelter ID æ•¸é‡: {num_unmatched}</p>", unsafe_allow_html=True)
    st.markdown(f"<p style='color:orange; font-weight:bold;'>ä¸Šå‚³æª”æ¡ˆä¸ç¬¦RMIçš„è³‡æ–™: {unmatched_data}</p>", unsafe_allow_html=True)
    st.markdown(f"<p style='color:green; font-weight:bold;'>ä¸Šå‚³æª”æ¡ˆç¬¦åˆRMIçš„ Smelter ID æ•¸é‡: {num_matched}</p>", unsafe_allow_html=True)
    st.subheader("è‡ªå‹•ç”ŸæˆOI")
    st.write(f"èˆ‡RMIä¸ç¬¦çš„è³‡æ–™ä¿å­˜æ–¼ï¼š[{unmatched_path}](file://{unmatched_path})")
    st.write(f"å«ä¾†æºåç¨±çš„ General æª”æ¡ˆä¿å­˜æ–¼ï¼š[{output_general_path}](file://{output_general_path})")
    st.write(f"Winbond å…¬ç‰ˆä¿å­˜æ–¼ï¼š[{winbond_path}](file://{winbond_path})")



def compare_versions(version_1, version_2, general_path, st):
    version_1_df = pd.read_excel(os.path.join(general_path, version_1))
    version_2_df = pd.read_excel(os.path.join(general_path, version_2))

    version_1_ids = set(version_1_df['Smelter Identification Number Input Column'])
    version_2_ids = set(version_2_df['Smelter Identification Number Input Column'])

    changes = {'æ–°å¢': version_2_ids - version_1_ids, 'ç§»é™¤': version_1_ids - version_2_ids}

    if changes['æ–°å¢']:
        added_data = []
        st.write("ç‰ˆæœ¬2æ–°å¢çš„ Smelter ID:")
        for smelter_id in changes['æ–°å¢']:
            metal = version_2_df.loc[version_2_df['Smelter Identification Number Input Column'] == smelter_id, 'Metal (*)'].values[0]
            source_name = version_2_df.loc[version_2_df['Smelter Identification Number Input Column'] == smelter_id, 'Source Name'].values[0]
            added_data.append({"Smelter ID": smelter_id, "Metal": metal, "Source Name": source_name})
        added_df = pd.DataFrame(added_data)
        st.table(added_df)

    if changes['ç§»é™¤']:
        st.write("ç‰ˆæœ¬2ç§»é™¤çš„ Smelter ID:")
        removed_data = []
        for smelter_id in changes['ç§»é™¤']:
            metal = version_1_df.loc[version_1_df['Smelter Identification Number Input Column'] == smelter_id, 'Metal (*)'].values[0]
            source_name = version_1_df.loc[version_1_df['Smelter Identification Number Input Column'] == smelter_id, 'Source Name'].values[0]
            removed_data.append({"Smelter ID": smelter_id, "Metal": metal, "Source Name": source_name})
    
    removed_df = pd.DataFrame(removed_data)
    st.table(removed_df)
    if not changes['æ–°å¢'] and not changes['ç§»é™¤']:
        st.write("æ²’æœ‰ Smelter ID çš„è®ŠåŒ–ã€‚")

def find_smelter_id(smelter_id_to_find, rmi_df, merge_df, st):
    smelter_id_to_find = smelter_id_to_find.strip()
    if not smelter_id_to_find:
        st.warning("è«‹è¼¸å…¥ Smelter IDï¼")
        return
    print(rmi_df)
    rmi_match = rmi_df[rmi_df['Smelter ID'] == smelter_id_to_find]
    merge_match = merge_df[merge_df['Smelter Identification Number Input Column'] == smelter_id_to_find]

    if not rmi_match.empty:
        st.write(f"Smelter ID {smelter_id_to_find} ç¬¦åˆRMI")
        st.dataframe(rmi_match)
    else:
        st.write(f"Smelter ID {smelter_id_to_find} ä¸ç¬¦åˆRMI")

    if not merge_match.empty:
        st.write(f"åœ¨ä¾›æ‡‰å•†/Subconæª”æ¡ˆä¸­æ‰¾åˆ° Smelter ID {smelter_id_to_find} çš„è³‡æ–™")
        st.dataframe(merge_match)
    else:
        st.write(f"åœ¨ä¾›æ‡‰å•†/Subconæª”æ¡ˆä¸­æœªæ‰¾åˆ° Smelter ID {smelter_id_to_find} çš„è³‡æ–™")
def compare_mineral_sources(compared_path, today_date):
    st.header("æ¯”å°è¯é‚¦èˆ‡RMIç¤¦ç”¢åœ°")
    st.markdown("<hr style='border: 1px solid lightgray;'>", unsafe_allow_html=True)
    rmi_file_path = st.session_state.get('rmi_file_path')
    merge_file_path = st.session_state.get('merged_file_path')

    if rmi_file_path and merge_file_path:
        if 'unmatched_rows' not in st.session_state or 'matched_rows' not in st.session_state:
            rmi_df = pd.read_excel(rmi_file_path)
            merge_df = pd.read_excel(merge_file_path)

            # é€²è¡Œæ¯”å°ä¸¦è™•ç†çš„å‡½æ•¸èª¿ç”¨
            unmatched_rows, matched_rows, unmatched_path, matched_path, result_text, num_unmatched, unmatched_data, num_matched = process_smelter_data(rmi_df, merge_df, compared_path, today_date)

            # å‰µå»ºå„ç‰ˆæœ¬çš„ Excel æ–‡ä»¶
            output_general_path, winbond_path = create_excel_files(merge_df, compared_path, today_date)

            # ä¿å­˜çµæœåˆ° session state
            st.session_state['unmatched_rows'] = unmatched_rows
            st.session_state['matched_rows'] = matched_rows
            st.session_state['unmatched_path'] = unmatched_path
            st.session_state['matched_path'] = matched_path
            st.session_state['result_text'] = result_text
            st.session_state['rmi_df'] = rmi_df
            st.session_state['merge_df'] = merge_df
            st.session_state['num_unmatched'] = num_unmatched
            st.session_state['unmatched_data'] = unmatched_data  # ç¢ºä¿å­˜å„² unmatched_data
            st.session_state['num_matched'] = num_matched
            st.session_state['output_general_path'] = output_general_path
            st.session_state['winbond_path'] = winbond_path  # é€™è£¡æ‰å­˜å„² winbond_path

        else:
            unmatched_rows = st.session_state['unmatched_rows']
            matched_rows = st.session_state['matched_rows']
            unmatched_path = st.session_state['unmatched_path']
            matched_path = st.session_state['matched_path']
            result_text = st.session_state['result_text']
            rmi_df = st.session_state['rmi_df']
            merge_df = st.session_state['merge_df']
            output_general_path = st.session_state['output_general_path']
            num_unmatched = st.session_state['num_unmatched']
            unmatched_data = st.session_state['unmatched_data']  # ç¢ºä¿è®€å– unmatched_data
            num_matched = st.session_state['num_matched']
            winbond_path = st.session_state['winbond_path']


        # é¡¯ç¤ºçµæœ
        display_results(num_unmatched, unmatched_path, output_general_path, winbond_path, result_text, num_matched, unmatched_data)

        st.subheader("è¼¸å…¥ Smelter ID é€²è¡ŒæŸ¥æ‰¾")
        smelter_id_to_find = st.text_input( '',placeholder="è«‹è¼¸å…¥å®Œæ•´çš„ Smelter IDï¼Œä¾‹å¦‚ï¼šCID001149")
        if st.button("æŸ¥æ‰¾"):
            find_smelter_id(smelter_id_to_find, rmi_df, merge_df, st)
    else:
        st.error("è«‹å…ˆåœ¨ 'çˆ¬å–RMI/åˆä½µè¯é‚¦æª”æ¡ˆ' é é¢è™•ç†æª”æ¡ˆã€‚")

            
def compare_general_versions():
    st.header("æ­·å²ç´€éŒ„æ¯”è¼ƒ")
    general_path = os.path.join(base_path, "merged")
    
    # åªé¸æ“‡ä»¥ "General" é–‹é ­çš„æª”æ¡ˆ
    version_files = [f for f in os.listdir(general_path) if f.startswith("General") and f.endswith(".xlsx")]

    if len(version_files) < 2:
        st.warning("è«‹ç¢ºä¿è‡³å°‘æœ‰å…©å€‹ General ç‰ˆæœ¬çš„æª”æ¡ˆ")
    else:
        version_1 = st.selectbox("é¸æ“‡ General ç‰ˆæœ¬ 1", version_files)
        version_2 = st.selectbox("é¸æ“‡ General ç‰ˆæœ¬ 2", version_files)

        if st.button("æ¯”è¼ƒ"):
            compare_versions(version_1, version_2, general_path, st)

def main():
    logo_path = "winbond.png"  # æ›¿æ›æˆä½ çš„ logo åœ–ç‰‡è·¯å¾‘
    col1, col2 = st.columns([1.8, 6])  # èª¿æ•´åˆ—çš„å¯¬åº¦

    with col1:
        # ä½¿ç”¨ PIL æ‰“é–‹åœ–ç‰‡ï¼Œç¢ºä¿åœ–ç‰‡å“è³ª
        logo = Image.open(logo_path)
        st.image(logo, use_column_width=True) # ä½¿ç”¨ st.image() ä¾†é¡¯ç¤º logo
    st.markdown("<div style='margin-top: 20px;'></div>", unsafe_allow_html=True)
    st.markdown("<h2 style='text-align: left;'>ğŸ” è¡çªç¤¦ç”¢æ¯”å°æŸ¥è©¢å¹³å°</h2>", unsafe_allow_html=True)
    st.markdown('<h4 style="color:#4a4a4a;">è«‹å…ˆè¨­ç½®ä¸€å€‹ç³»çµ±è³‡æ–™å¤¾ï¼Œä»¥æ”¾ç½®æœ¬ç³»çµ±ç”Ÿæˆä¹‹æª”æ¡ˆ</h4>', unsafe_allow_html=True)
    st.markdown("""
        åœ¨ç³»çµ±è³‡æ–™å¤¾ï¼Œæ”¾ç½®ä»¥ä¸‹è³‡æ–™ï¼š
        - **Chrome Driver**
            - [ä¸‹è¼‰é€£çµ](https://storage.googleapis.com/chrome-for-testing-public/127.0.6533.119/win64/chromedriver-win64.zip)
        - **RMI æä¾› CMRT ç©ºè¡¨æ ¼**
    """)
    
    # è·¯å¾‘è¨­ç½®åŠŸèƒ½
    base_path, today_date, compared_path = setup_paths()
    if not base_path:
        st.info("è¼¸å…¥å¾Œï¼Œè«‹æŒ‰Enteré€å‡ºè·¯å¾‘ã€‚")
        return

    # åŠ è¼‰è‡ªå®šç¾© CSS
    st.markdown('<style>' + open('./style.css').read() + '</style>', unsafe_allow_html=True)

    with st.sidebar:
        tabs = on_hover_tabs(tabName=['çˆ¬å–RMI/åˆä½µè¯é‚¦æª”æ¡ˆ', 'æ¯”å°è¯é‚¦èˆ‡RMIç¤¦ç”¢åœ°', 'æ­·å²ç´€éŒ„æ¯”è¼ƒ'],
                             iconName=['file_download', 'compare_arrows', 'assessment'],
                             default_choice=0)

    scroll_script = """
        <script>
            function scrollToBottom() {
                window.scrollTo({top: document.body.scrollHeight, behavior: 'smooth'});
            }
            scrollToBottom();
        </script>
    """
    
    if tabs == 'çˆ¬å–RMI/åˆä½µè¯é‚¦æª”æ¡ˆ':
        st.markdown(scroll_script, unsafe_allow_html=True)
        st.markdown("<div style='margin-top: 50px;'></div>", unsafe_allow_html=True)  
        st.markdown("<div style='margin-down: 100px;'></div>", unsafe_allow_html=True)
        download_and_merge_files(base_path, today_date)
    elif tabs == 'æ¯”å°è¯é‚¦èˆ‡RMIç¤¦ç”¢åœ°':
        st.markdown(scroll_script, unsafe_allow_html=True)
        st.markdown("<div style='margin-top: 50px;'></div>", unsafe_allow_html=True)  
        st.markdown("<div style='margin-down: 100px;'></div>", unsafe_allow_html=True)
        compare_mineral_sources(compared_path, today_date)
    elif tabs == 'æ­·å²ç´€éŒ„æ¯”è¼ƒ':
        st.markdown(scroll_script, unsafe_allow_html=True)
        st.markdown("<div style='margin-top: 50px;'></div>", unsafe_allow_html=True)  
        st.markdown("<div style='margin-down: 100px;'></div>", unsafe_allow_html=True)
        compare_general_versions()

if __name__ == "__main__":
    main()
